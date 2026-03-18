import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import io

# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="RK Ops Monitor",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: #0a0a0a; color: #f0f0f0; }
.main-title { font-family:'Syne',sans-serif; font-size:2rem; font-weight:800; color:#fff; letter-spacing:-1px; margin-bottom:0; }
.main-sub   { color:#444; font-size:0.85rem; margin-bottom:1.5rem; }
.section-label { font-family:'Syne',sans-serif; font-size:0.72rem; font-weight:700; color:#e8a020; letter-spacing:2px; text-transform:uppercase; margin-bottom:0.4rem; }
.card { background:#111; border:1px solid #1e1e1e; border-radius:10px; padding:1rem 1.2rem; margin-bottom:0.8rem; }
.metric-up   { color:#4ade80; font-weight:700; }
.metric-down { color:#f87171; font-weight:700; }
.metric-flat { color:#888; }
.chip-red    { background:#3f0f0f; color:#f87171; border-radius:5px; padding:2px 8px; font-size:0.78rem; display:inline-block; }
.chip-green  { background:#0f2f1f; color:#4ade80; border-radius:5px; padding:2px 8px; font-size:0.78rem; display:inline-block; }
.chip-amber  { background:#2f2000; color:#fbbf24; border-radius:5px; padding:2px 8px; font-size:0.78rem; display:inline-block; }
.step-badge  { background:#e8a020; color:#000; border-radius:50%; width:20px; height:20px; display:inline-block; text-align:center; line-height:20px; font-weight:800; font-size:0.75rem; margin-right:8px; }
hr.div { border:none; border-top:1px solid #1e1e1e; margin:1.2rem 0; }
.stButton > button { background:#e8a020 !important; color:#000 !important; font-family:'Syne',sans-serif !important; font-weight:700 !important; border-radius:8px !important; border:none !important; }
.stButton > button:hover { background:#f5b535 !important; }
div[data-testid="stExpander"] { background:#111 !important; border:1px solid #1e1e1e !important; border-radius:8px !important; }
.stTabs [data-baseweb="tab-list"] { background:#111 !important; border-radius:8px !important; gap:4px; }
.stTabs [aria-selected="true"] { color:#e8a020 !important; }
.stTabs [data-baseweb="tab"] { color:#555 !important; }
.stDataFrame { background:#111; }
div[data-testid="metric-container"] { background:#111; border:1px solid #1e1e1e; border-radius:8px; padding:0.8rem; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HARDCODED MAPPING — TL → Outlets (same IDs as monthly app)
# ══════════════════════════════════════════════════════════════════════════════
CITY_MAPPING = {
    "Navneet Singh (Noida)": [
        {"outlet":"Sec 104",           "pos":28039,  "zmt_rk":19476740,"swg_rk":313666},
        {"outlet":"Sector-141 Noida",  "pos":26592,  "zmt_rk":18734595,"swg_rk":63465},
        {"outlet":"Sector-132 Noida",  "pos":34303,  "zmt_rk":18750756,"swg_rk":68184},
        {"outlet":"Sector 125 Noida",  "pos":373602, "zmt_rk":21824279,"swg_rk":1069514},
        {"outlet":"Sector-73 Noida",   "pos":97074,  "zmt_rk":20508934,"swg_rk":635149},
        {"outlet":"Sector-44 Noida",   "pos":39966,  "zmt_rk":18575970,"swg_rk":42813},
    ],
    "Ajay Halder (Noida/Gzb)": [
        {"outlet":"Sector 4 Noida",    "pos":21787,  "zmt_rk":19364731,"swg_rk":54622},
        {"outlet":"Sector-62",         "pos":23687,  "zmt_rk":302308,  "swg_rk":42808},
        {"outlet":"Sector-37",         "pos":32851,  "zmt_rk":20374787,"swg_rk":583789},
        {"outlet":"Sector-18",         "pos":26952,  "zmt_rk":304612,  "swg_rk":42807},
        {"outlet":"Gaur City GNoida",  "pos":112772, "zmt_rk":20589872,"swg_rk":879431},
        {"outlet":"Eco Loft",          "pos":74178,  "zmt_rk":20264919,"swg_rk":531120},
    ],
    "Sunil Sharma (Ghaziabad)": [
        {"outlet":"RDC Raj Nagar Gzb", "pos":363143, "zmt_rk":18962941,"swg_rk":879460},
        {"outlet":"GNB Mall",          "pos":113953, "zmt_rk":21341669,"swg_rk":1082869},
        {"outlet":"Shipra Mall",       "pos":408910, "zmt_rk":22103426,"swg_rk":1238359},
    ],
    "Vishwanath Rao (Ghaziabad)": [
        {"outlet":"Indirapuram",       "pos":38041,  "zmt_rk":18633334,"swg_rk":46674},
        {"outlet":"Rajendra Nagar Gzb","pos":37055,  "zmt_rk":19283683,"swg_rk":241917},
        {"outlet":"Vasundhra",         "pos":122466, "zmt_rk":20711593,"swg_rk":731841},
    ],
    "Sanjay Morya (Delhi/Gurugram)": [
        {"outlet":"Kalkaji",           "pos":31247,  "zmt_rk":18869459,"swg_rk":90719},
        {"outlet":"Tilak Nagar",       "pos":63819,  "zmt_rk":18942689,"swg_rk":123197},
        {"outlet":"Vasant Kunj",       "pos":25924,  "zmt_rk":19030978,"swg_rk":131217},
        {"outlet":"Chattarpur",        "pos":26423,  "zmt_rk":19052007,"swg_rk":140433},
        {"outlet":"Paschim Vihar",     "pos":43412,  "zmt_rk":20256577,"swg_rk":531480},
        {"outlet":"Gtb Nagar",         "pos":79050,  "zmt_rk":20323930,"swg_rk":569414},
        {"outlet":"Nathupur Gurugram", "pos":108782, "zmt_rk":20582763,"swg_rk":668475},
        {"outlet":"Old DLF Gurugram",  "pos":108777, "zmt_rk":20582827,"swg_rk":668470},
        {"outlet":"Sector-57 Gurugram","pos":74068,  "zmt_rk":20463325,"swg_rk":624165},
        {"outlet":"Wazirabad Gurugram","pos":108779, "zmt_rk":20582847,"swg_rk":668467},
        {"outlet":"Gurugram Sec-82",   "pos":30407,  "zmt_rk":19513923,"swg_rk":327106},
        {"outlet":"Sector 90 Gurugram","pos":380769, "zmt_rk":21929020,"swg_rk":1102249},
        {"outlet":"Rohini",            "pos":93493,  "zmt_rk":22100897,"swg_rk":622353},
        {"outlet":"Vikashpuri",        "pos":404584, "zmt_rk":22227640,"swg_rk":1224350},
        {"outlet":"Uttam Nagar Dwarka","pos":414828, "zmt_rk":None,    "swg_rk":1280807},
        {"outlet":"Subhash Nagar",     "pos":398993, "zmt_rk":22165860,"swg_rk":1196325},
    ],
    "Zeeshan Ali (Faridabad)": [
        {"outlet":"Shaheen Bagh",      "pos":118685, "zmt_rk":20666436,"swg_rk":704360},
        {"outlet":"NIT Faridabad",     "pos":96843,  "zmt_rk":20480333,"swg_rk":632083},
        {"outlet":"Sec-15 Faridabad",  "pos":54369,  "zmt_rk":18567324,"swg_rk":42815},
        {"outlet":"Lakkarpur Faridabad","pos":143500,"zmt_rk":20873208,"swg_rk":775707},
        {"outlet":"Greenfield Faridabad","pos":154254,"zmt_rk":21446399,"swg_rk":983943},
    ],
    "Badir Alam (MP)": [
        {"outlet":"Bhopal",            "pos":338959, "zmt_rk":21340655,"swg_rk":934354},
        {"outlet":"Indore",            "pos":109589, "zmt_rk":20566161,"swg_rk":673809},
        {"outlet":"Siddharth Nagar Indore","pos":156653,"zmt_rk":21022031,"swg_rk":690867},
    ],
    "Abhishek Kumar (Bangalore)": [
        {"outlet":"Whitefield",        "pos":89397,  "zmt_rk":20410563,"swg_rk":606509},
        {"outlet":"Mahadevpura",       "pos":72269,  "zmt_rk":20201048,"swg_rk":515199},
        {"outlet":"Koramangala",       "pos":83769,  "zmt_rk":20359621,"swg_rk":580691},
        {"outlet":"Electronic City",   "pos":72413,  "zmt_rk":20213913,"swg_rk":515053},
        {"outlet":"Sarjapur",          "pos":68691,  "zmt_rk":20163232,"swg_rk":494751},
        {"outlet":"Kalyan Nagar",      "pos":75899,  "zmt_rk":20265149,"swg_rk":544214},
        {"outlet":"Bel Road",          "pos":75897,  "zmt_rk":20263151,"swg_rk":536015},
        {"outlet":"Habble Bangalore",  "pos":95682,  "zmt_rk":20471662,"swg_rk":625912},
        {"outlet":"Indira Nagar",      "pos":403199, "zmt_rk":22179137,"swg_rk":1203098},
    ],
    "Virendra Pratap (Chennai)": [
        {"outlet":"Mohanram Nagar",    "pos":84743,  "zmt_rk":20410863,"swg_rk":588878},
        {"outlet":"Madipakkam",        "pos":84742,  "zmt_rk":20410826,"swg_rk":588790},
        {"outlet":"Parengudi Chennai", "pos":97078,  "zmt_rk":20486896,"swg_rk":631195},
    ],
    "Atul Kumar (Pune)": [
        {"outlet":"Apple Ghar Pune",   "pos":129883, "zmt_rk":20748035,"swg_rk":733937},
        {"outlet":"Hinjewadi Phase 1", "pos":141096, "zmt_rk":20855724,"swg_rk":756772},
        {"outlet":"Millennium Mall",   "pos":137998, "zmt_rk":21067154,"swg_rk":833916},
        {"outlet":"Shivaji Nagar Pune","pos":346318, "zmt_rk":21435196,"swg_rk":354312},
    ],
    "Bhupesh Bhatt (Hyderabad)": [
        {"outlet":"Madhapur",          "pos":24485,  "zmt_rk":18953624,"swg_rk":120196},
        {"outlet":"Gachibowli",        "pos":24487,  "zmt_rk":19271816,"swg_rk":214621},
        {"outlet":"Banjara Hills",     "pos":129436, "zmt_rk":21028217,"swg_rk":711834},
        {"outlet":"Taranagar Hyderabad","pos":141099,"zmt_rk":20855101,"swg_rk":766665},
        {"outlet":"RK Puram Hyderabad","pos":44718,  "zmt_rk":19714313,"swg_rk":375980},
        {"outlet":"Lulu Mall Hyderabad","pos":141214,"zmt_rk":21154081,"swg_rk":866698},
        {"outlet":"Miyapur",           "pos":24489,  "zmt_rk":21779883,"swg_rk":1063096},
        {"outlet":"Goa Anjuna",        "pos":339817, "zmt_rk":21365117,"swg_rk":946493},
    ],
    "Milan (Mumbai)": [
        {"outlet":"G Corp",            "pos":367105, "zmt_rk":21734559,"swg_rk":1063584},
        {"outlet":"Mumbai Pawai",      "pos":353027, "zmt_rk":21522379,"swg_rk":985771},
        {"outlet":"Raymond",           "pos":369670, "zmt_rk":21794492,"swg_rk":1066710},
        {"outlet":"Mumbai BKC",        "pos":375018, "zmt_rk":21824216,"swg_rk":1076952},
        {"outlet":"Mumbai Chembur",    "pos":383109, "zmt_rk":21966993,"swg_rk":1104174},
        {"outlet":"Airoli Navi Mumbai","pos":386396, "zmt_rk":21982077,"swg_rk":1123441},
        {"outlet":"Mumbai Dahisar",    "pos":399741, "zmt_rk":22170342,"swg_rk":1220916},
        {"outlet":"Mumbai Marol",      "pos":404829, "zmt_rk":22274879,"swg_rk":1238360},
        {"outlet":"Mira Road Mumbai",  "pos":386734, "zmt_rk":22044961,"swg_rk":1142360},
    ],
}

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def safe_id(v):
    try:
        s = str(v).strip()
        return None if s in ('#N/A','','None','nan') else int(float(s))
    except:
        return None

def safe_f(v, d=0.0):
    try:
        f = float(v if v is not None else d)
        return d if f != f else f
    except:
        return d

def parse_pct(v):
    if v is None: return None
    try:
        s = str(v).replace('%','').strip()
        f = float(s)
        return f if f > 1 else f * 100
    except:
        return None

def parse_min(v):
    if v is None: return None
    try:
        return float(str(v).replace(' mins','').replace(' min','').strip())
    except:
        return None

def arrow(curr, prev, lower_is_better=False):
    """Return direction arrow and CSS class"""
    if curr is None or prev is None or prev == 0:
        return "—", "metric-flat"
    if curr > prev:
        cls = "metric-down" if lower_is_better else "metric-up"
        return "↑", cls
    elif curr < prev:
        cls = "metric-up" if lower_is_better else "metric-down"
        return "↓", cls
    return "→", "metric-flat"

def pct_change(curr, prev):
    if curr is None or prev is None or prev == 0:
        return None
    return round((curr - prev) / prev * 100, 1)

# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADERS — same format as monthly app
# ══════════════════════════════════════════════════════════════════════════════
def load_zomato(wb):
    """Load Zomato data → {rid: {orders, complaints, kpt, rating, online_pct}}"""
    zmt = {}
    for sname in wb.sheetnames:
        if 'zomato' not in sname.lower(): continue
        for row in wb[sname].iter_rows(min_row=2, values_only=True):
            rid = safe_id(row[0])
            if not rid: continue
            metric = str(row[5]).strip() if row[5] else ''
            val    = row[6]
            if rid not in zmt:
                zmt[rid] = {'orders':0,'complaints':0,'kpt':None,'rating':None,'online_pct':None}
            if   metric == 'Delivered orders':   zmt[rid]['orders']     = safe_f(val)
            elif metric == 'Total complaints':   zmt[rid]['complaints'] = safe_f(val)
            elif metric == 'KPT (in minutes)':   zmt[rid]['kpt']        = safe_f(val) if val else None
            elif metric == 'Average rating':     zmt[rid]['rating']     = safe_f(val) if val else None
            elif metric == 'Online %':           zmt[rid]['online_pct'] = parse_pct(val)
        break
    return zmt

def load_swiggy(wb):
    """Load Swiggy data → {rid: {orders, kpt, avail, cmp_pct, rating}}"""
    swg = {}
    for sname in wb.sheetnames:
        if 'swiggy' not in sname.lower(): continue
        for row in wb[sname].iter_rows(min_row=2, values_only=True):
            rid = safe_id(row[0])
            if not rid: continue
            metric = str(row[5]).strip() if row[5] else ''
            val    = row[6]
            if rid not in swg:
                swg[rid] = {'orders':0,'kpt':None,'avail':None,'cmp_pct':None,'rating':None}
            if   metric in ('Delivered Orders','Orders'):
                swg[rid]['orders']  = safe_f(val)
            elif metric == 'Avg Prep Time':
                v = parse_min(val)
                if v and v > 0: swg[rid]['kpt'] = v
            elif metric == 'Online Availability %':
                swg[rid]['avail']   = parse_pct(val)
            elif metric == '% Orders with Complaints':
                swg[rid]['cmp_pct'] = parse_pct(val)
            elif metric in ('Rating','Average Rating'):
                swg[rid]['rating']  = safe_f(val) if val else None
        break
    return swg

def detect_file(file_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheets = [s.lower() for s in wb.sheetnames]
        has_z = any('zomato' in s for s in sheets)
        has_s = any('swiggy' in s for s in sheets)
        if has_z or has_s:
            return True, wb
        return False, wb
    except:
        return False, None

# ══════════════════════════════════════════════════════════════════════════════
# CALCULATOR — per outlet metrics from raw data
# ══════════════════════════════════════════════════════════════════════════════
def calc_outlet_metrics(outlet, zmt, swg):
    """Return dict of metrics for one outlet"""
    z_rk = outlet.get('zmt_rk')
    s_rk = outlet.get('swg_rk')

    # Orders
    z_orders = zmt.get(z_rk, {}).get('orders', 0) if z_rk else 0
    s_orders = swg.get(s_rk, {}).get('orders', 0)  if s_rk else 0
    total_orders = z_orders + s_orders

    # Complaints — Zomato raw count + Swiggy back-calc from %
    z_comps = zmt.get(z_rk, {}).get('complaints', 0) if z_rk else 0
    s_cmp_pct = swg.get(s_rk, {}).get('cmp_pct') if s_rk else None
    s_comps = round(s_cmp_pct / 100 * s_orders) if (s_cmp_pct and s_orders > 0) else 0
    total_comps = z_comps + s_comps
    cmp_pct = round(total_comps / total_orders * 100, 2) if total_orders > 0 else None

    # KPT — Zomato RK only (confirmed from analysis)
    z_kpt = zmt.get(z_rk, {}).get('kpt') if z_rk else None
    kpt = round(z_kpt, 2) if (z_kpt and z_kpt > 0) else None

    # Rating — Zomato primary, Swiggy fallback
    z_rat = zmt.get(z_rk, {}).get('rating') if z_rk else None
    s_rat = swg.get(s_rk, {}).get('rating')  if s_rk else None
    rats = [r for r in [z_rat, s_rat] if r and r > 0]
    rating = round(sum(rats)/len(rats), 2) if rats else None
    rating_src = "Zomato+Swiggy" if len(rats)==2 else ("Zomato" if z_rat else ("Swiggy" if s_rat else "N/A"))

    # Availability
    z_avail = zmt.get(z_rk, {}).get('online_pct') if z_rk else None
    s_avail = swg.get(s_rk, {}).get('avail')       if s_rk else None
    avails  = [a for a in [z_avail, s_avail] if a is not None]
    avail   = round(sum(avails)/len(avails), 2) if avails else None

    return {
        'orders':       total_orders,
        'z_orders':     z_orders,
        's_orders':     s_orders,
        'total_comps':  total_comps,
        'cmp_pct':      cmp_pct,
        'kpt':          kpt,
        'kpt_breach':   (1 if kpt and kpt > 12 else 0) if kpt is not None else None,
        'rating':       rating,
        'rating_src':   rating_src,
        'avail':        avail,
        'has_data':     (total_orders > 0 or kpt is not None or rating is not None),
    }

def calc_city_summary(city, outlets, zmt, swg):
    """Aggregate metrics across all outlets in a city"""
    rows = []
    for o in outlets:
        m = calc_outlet_metrics(o, zmt, swg)
        m['outlet'] = o['outlet']
        rows.append(m)

    active = [r for r in rows if r['has_data']]
    if not active:
        return None, rows

    total_orders = sum(r['orders'] for r in active)
    total_comps  = sum(r['total_comps'] for r in active)
    cmp_pct      = round(total_comps / total_orders * 100, 2) if total_orders > 0 else None

    kpt_vals     = [r['kpt'] for r in active if r['kpt'] is not None]
    avg_kpt      = round(sum(kpt_vals)/len(kpt_vals), 2) if kpt_vals else None
    kpt_breaches = sum(1 for r in active if r['kpt'] and r['kpt'] > 12)

    rat_vals     = [r['rating'] for r in active if r['rating'] is not None]
    avg_rating   = round(sum(rat_vals)/len(rat_vals), 2) if rat_vals else None

    return {
        'city':         city,
        'outlets':      len(active),
        'total_orders': total_orders,
        'total_comps':  total_comps,
        'cmp_pct':      cmp_pct,
        'avg_kpt':      avg_kpt,
        'kpt_breaches': kpt_breaches,
        'avg_rating':   avg_rating,
    }, rows

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════
def thin_border():
    s = Side(style="thin", color="2a2a2a")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, val, bg="1a1a1a", fg="e8a020", sz=9, bold=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, color=fg, size=sz, name="Arial")
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()

def data_cell(ws, row, col, val, bg="0f0f0f", fg="f0f0f0", sz=9, bold=False, center=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, color=fg, size=sz, name="Arial")
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    c.border    = thin_border()

def build_excel_report(summaries_curr, summaries_prev, outlet_detail,
                        report_type, label_curr, label_prev):
    wb = openpyxl.Workbook()

    # ── Sheet 1: City Summary ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "City Summary"
    ws1.sheet_view.showGridLines = False

    title = f"RollsKing — {report_type} Operations Report | {label_curr} vs {label_prev}"
    ws1.merge_cells("A1:O1")
    c = ws1["A1"]
    c.value     = title
    c.font      = Font(bold=True, size=13, color="e8a020", name="Arial")
    c.fill      = PatternFill("solid", start_color="0a0a0a")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:O2")
    c = ws1["A2"]
    c.value     = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}  |  Current: {label_curr}  |  Previous: {label_prev}"
    c.font      = Font(size=8, color="555555", italic=True, name="Arial")
    c.fill      = PatternFill("solid", start_color="0f0f0f")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 16

    # Group headers
    hdr_groups = [
        ("A3:A4", "City / Manager"),
        ("B3:B4", "Outlets"),
        ("C3:E3", "Orders"),
        ("F3:H3", "Complaints"),
        ("I3:K3", "KPT (min)"),
        ("L3:N3", "Rating"),
    ]
    for span, text in hdr_groups:
        ws1.merge_cells(span)
        c = ws1[span.split(":")[0]]
        c.value     = text
        c.font      = Font(bold=True, color="e8a020", size=9, name="Arial")
        c.fill      = PatternFill("solid", start_color="1a1a1a")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

    sub_hdrs = {3: label_curr, 4: label_prev, 5: "Δ%",
                6: label_curr, 7: label_prev, 8: "Δ (pp)",
                9: label_curr, 10: label_prev, 11: "Breaches",
                12: label_curr, 13: label_prev, 14: "Δ"}
    for i in range(3, 15):  # skip cols 1,2 — already merged from row 3
        h = sub_hdrs.get(i, "")
        hdr_cell(ws1, 4, i, h, bg="141414", fg="aaaaaa", sz=8)
    ws1.row_dimensions[3].height = 18
    ws1.row_dimensions[4].height = 16

    col_w = [28,8,12,12,10, 10,10,10, 10,10,10, 10,10,10]
    for i, w in enumerate(col_w, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    GOLD  = "e8a020"
    RED   = "f87171"
    GREEN = "4ade80"
    DARK  = "0f0f0f"
    ALT   = "111111"

    for i, city in enumerate(summaries_curr.keys()):
        row  = i + 5
        curr = summaries_curr.get(city)
        prev = summaries_prev.get(city)
        bg   = ALT if i % 2 else DARK

        def val(d, k): return d.get(k) if d else None

        # Orders
        c_ord = val(curr,'total_orders'); p_ord = val(prev,'total_orders')
        d_ord = pct_change(c_ord, p_ord)

        # Complaints
        c_cmp = val(curr,'cmp_pct'); p_cmp = val(prev,'cmp_pct')
        d_cmp = round(c_cmp - p_cmp, 2) if (c_cmp and p_cmp) else None

        # KPT
        c_kpt = val(curr,'avg_kpt'); p_kpt = val(prev,'avg_kpt')
        c_kpb = val(curr,'kpt_breaches')

        # Rating
        c_rat = val(curr,'avg_rating'); p_rat = val(prev,'avg_rating')
        d_rat = round(c_rat - p_rat, 2) if (c_rat and p_rat) else None

        data_cell(ws1, row, 1,  city,  bg=bg, fg="ffffff", bold=True, center=False)
        data_cell(ws1, row, 2,  val(curr,'outlets') or "—", bg=bg)
        data_cell(ws1, row, 3,  c_ord, bg=bg)
        data_cell(ws1, row, 4,  p_ord, bg=bg, fg="555555")
        fg_d = GREEN if (d_ord and d_ord > 0) else (RED if d_ord and d_ord < 0 else "555555")
        data_cell(ws1, row, 5,  f"{d_ord:+.1f}%" if d_ord is not None else "—", bg=bg, fg=fg_d)

        # Complaints — red if high
        fg_c = RED if (c_cmp and c_cmp >= 3) else (GREEN if c_cmp and c_cmp < 2 else "f0f0f0")
        data_cell(ws1, row, 6,  f"{c_cmp:.2f}%" if c_cmp else "—", bg=bg, fg=fg_c)
        data_cell(ws1, row, 7,  f"{p_cmp:.2f}%" if p_cmp else "—", bg=bg, fg="555555")
        fg_dc = GREEN if (d_cmp and d_cmp < 0) else (RED if d_cmp and d_cmp > 0 else "555555")
        data_cell(ws1, row, 8,  f"{d_cmp:+.2f}" if d_cmp is not None else "—", bg=bg, fg=fg_dc)

        # KPT — red if > 12
        fg_k = RED if (c_kpt and c_kpt > 12) else (GREEN if c_kpt and c_kpt <= 10 else "f0f0f0")
        data_cell(ws1, row, 9,  f"{c_kpt:.1f}" if c_kpt else "—", bg=bg, fg=fg_k)
        data_cell(ws1, row, 10, f"{p_kpt:.1f}" if p_kpt else "—", bg=bg, fg="555555")
        fg_kb = RED if (c_kpb and c_kpb > 0) else GREEN
        data_cell(ws1, row, 11, c_kpb if c_kpb is not None else "—", bg=bg, fg=fg_kb)

        # Rating — red if < 4
        fg_r = RED if (c_rat and c_rat < 4.0) else (GREEN if c_rat and c_rat >= 4.3 else "f0f0f0")
        data_cell(ws1, row, 12, f"{c_rat:.2f}" if c_rat else "—", bg=bg, fg=fg_r)
        data_cell(ws1, row, 13, f"{p_rat:.2f}" if p_rat else "—", bg=bg, fg="555555")
        fg_dr = GREEN if (d_rat and d_rat > 0) else (RED if d_rat and d_rat < 0 else "555555")
        data_cell(ws1, row, 14, f"{d_rat:+.2f}" if d_rat is not None else "—", bg=bg, fg=fg_dr)
        ws1.row_dimensions[row].height = 16

    # ── Sheet 2: Outlet Detail ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Outlet Detail")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:L1")
    c = ws2["A1"]
    c.value     = f"Outlet Detail — {label_curr} vs {label_prev}"
    c.font      = Font(bold=True, size=12, color="e8a020", name="Arial")
    c.fill      = PatternFill("solid", start_color="0a0a0a")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    od_hdrs = ["City/Manager","Outlet","Orders","Complaints","Cmp %","KPT (min)","KPT Breach","Rating","Avail %","vs Prev Orders","vs Prev Cmp%","vs Prev Rating"]
    for i, h in enumerate(od_hdrs, 1):
        hdr_cell(ws2, 2, i, h)
    ws2.row_dimensions[2].height = 20

    for col_l, w in zip([get_column_letter(i) for i in range(1,13)],
                         [24,22,10,12,10,10,10,10,10,12,12,12]):
        ws2.column_dimensions[col_l].width = w

    row_num = 3
    for city, rows in outlet_detail.items():
        prev_rows = {r['outlet']: r for r in outlet_detail.get(f"_prev_{city}", [])}
        for r in rows:
            bg = ALT if row_num % 2 else DARK
            pr = prev_rows.get(r['outlet'], {})
            d_ord = pct_change(r.get('orders'), pr.get('orders'))
            d_cmp = round(r['cmp_pct'] - pr['cmp_pct'], 2) if (r.get('cmp_pct') and pr.get('cmp_pct')) else None
            d_rat = round(r['rating'] - pr['rating'], 2) if (r.get('rating') and pr.get('rating')) else None

            vals = [
                city, r['outlet'],
                r.get('orders') or "—",
                r.get('total_comps') or "—",
                f"{r['cmp_pct']:.2f}%" if r.get('cmp_pct') else "—",
                f"{r['kpt']:.1f}" if r.get('kpt') else "—",
                "YES" if r.get('kpt_breach') else ("—" if r.get('kpt') is None else "OK"),
                f"{r['rating']:.2f}" if r.get('rating') else "—",
                f"{r['avail']:.1f}%" if r.get('avail') else "—",
                f"{d_ord:+.1f}%" if d_ord is not None else "—",
                f"{d_cmp:+.2f}" if d_cmp is not None else "—",
                f"{d_rat:+.2f}" if d_rat is not None else "—",
            ]
            for col, val in enumerate(vals, 1):
                fg = "ffffff"
                if col == 7 and val == "YES": fg = RED
                if col == 7 and val == "OK":  fg = GREEN
                if col == 8 and r.get('rating') and r['rating'] < 4.0: fg = RED
                data_cell(ws2, row_num, col, val, bg=bg, fg=fg,
                          center=(col != 2), bold=(col == 2))
            ws2.row_dimensions[row_num].height = 15
            row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# PDF BUILDER
# ══════════════════════════════════════════════════════════════════════════════
def build_pdf_report(summaries_curr, summaries_prev, outlet_detail,
                      report_type, label_curr, label_prev):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np
    plt.close('all')

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, PageBreak, HRFlowable)
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    C_BG   = "#0a0a0a"; C_PANEL = "#111111"; C_GOLD = "#e8a020"
    C_TEXT = "#f0f0f0"; C_SUB   = "#666666"
    C_RED  = "#f87171"; C_GREEN = "#4ade80"; C_AMBER = "#fbbf24"

    def fig_to_rl(fig, w_mm, h_mm):
        b = io.BytesIO()
        fig.savefig(b, format='png', dpi=150, bbox_inches='tight',
                    facecolor=fig.get_facecolor())
        b.seek(0); plt.close(fig)
        return RLImage(b, width=w_mm*mm, height=h_mm*mm)

    def chart_complaints_kpt():
        cities = list(summaries_curr.keys())
        curr_cmps = [summaries_curr[c].get('cmp_pct') or 0 for c in cities]
        prev_cmps = [summaries_prev.get(c, {}).get('cmp_pct') or 0 for c in cities]
        curr_kpts = [summaries_curr[c].get('avg_kpt') or 0 for c in cities]
        prev_kpts = [summaries_prev.get(c, {}).get('avg_kpt') or 0 for c in cities]

        short = [c.split('(')[0].strip()[:12] for c in cities]
        x = np.arange(len(cities))
        w = 0.35

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5.5))
        fig.patch.set_facecolor(C_BG)

        for ax in (ax1, ax2):
            ax.set_facecolor(C_PANEL)
            [ax.spines[s].set_color('#333') for s in ax.spines]
            ax.tick_params(colors=C_SUB, labelsize=7.5)
            ax.set_axisbelow(True)
            ax.yaxis.grid(True, color='#1e1e1e', linewidth=0.5)

        # Complaints chart
        bars1 = ax1.bar(x - w/2, prev_cmps, w, color='#333344', label=label_prev, alpha=0.9)
        bars2 = ax1.bar(x + w/2, curr_cmps, w, color=C_GOLD, label=label_curr, alpha=0.9)
        ax1.axhline(y=3, color=C_RED, linewidth=1.5, linestyle='--', alpha=0.7)
        ax1.axhline(y=1, color=C_GREEN, linewidth=1, linestyle=':', alpha=0.5)
        for bar in bars2:
            h = bar.get_height()
            if h > 0:
                clr = C_RED if h >= 3 else (C_AMBER if h >= 2 else C_GREEN)
                bar.set_color(clr)
        ax1.set_xticks(x); ax1.set_xticklabels(short, rotation=30, ha='right', fontsize=7)
        ax1.set_ylabel("Complaint %", color=C_SUB, fontsize=8)
        ax1.set_title("Complaints % — Current vs Previous", color=C_TEXT, fontsize=10, fontweight='bold', pad=8)
        ax1.legend(fontsize=7, facecolor=C_PANEL, edgecolor='#333', labelcolor=C_TEXT)
        ax1.set_ylim(0, max(max(curr_cmps + prev_cmps) * 1.2, 4))

        # KPT chart
        bars3 = ax2.bar(x - w/2, prev_kpts, w, color='#333344', label=label_prev, alpha=0.9)
        bars4 = ax2.bar(x + w/2, curr_kpts, w, color=C_GOLD, label=label_curr, alpha=0.9)
        ax2.axhline(y=12, color=C_RED, linewidth=1.5, linestyle='--', alpha=0.7)
        for bar in bars4:
            h = bar.get_height()
            if h > 12: bar.set_color(C_RED)
            elif h > 0: bar.set_color(C_GREEN)
        ax2.set_xticks(x); ax2.set_xticklabels(short, rotation=30, ha='right', fontsize=7)
        ax2.set_ylabel("Avg KPT (min)", color=C_SUB, fontsize=8)
        ax2.set_title("KPT — Current vs Previous", color=C_TEXT, fontsize=10, fontweight='bold', pad=8)
        ax2.legend(fontsize=7, facecolor=C_PANEL, edgecolor='#333', labelcolor=C_TEXT)
        ax2.set_ylim(0, max(max(curr_kpts + prev_kpts) * 1.2, 14))

        fig.tight_layout(pad=1.5)
        return fig_to_rl(fig, 175, 80)

    def chart_ratings():
        cities = [c for c in summaries_curr if summaries_curr[c].get('avg_rating')]
        if not cities: return None
        curr_rats = [summaries_curr[c]['avg_rating'] for c in cities]
        prev_rats = [summaries_prev.get(c, {}).get('avg_rating') or 0 for c in cities]
        short = [c.split('(')[0].strip()[:12] for c in cities]
        x = np.arange(len(cities)); w = 0.35

        fig, ax = plt.subplots(figsize=(14, 4))
        fig.patch.set_facecolor(C_BG); ax.set_facecolor(C_PANEL)
        [ax.spines[s].set_color('#333') for s in ax.spines]
        ax.tick_params(colors=C_SUB, labelsize=7.5)
        ax.set_axisbelow(True); ax.yaxis.grid(True, color='#1e1e1e', linewidth=0.5)

        ax.bar(x - w/2, prev_rats, w, color='#333344', label=label_prev, alpha=0.9)
        bars = ax.bar(x + w/2, curr_rats, w, color=C_GOLD, label=label_curr, alpha=0.9)
        for bar, v in zip(bars, curr_rats):
            bar.set_color(C_RED if v < 4.0 else (C_GREEN if v >= 4.3 else C_GOLD))
        ax.axhline(y=4.0, color=C_RED, linewidth=1.5, linestyle='--', alpha=0.7)
        ax.set_xticks(x); ax.set_xticklabels(short, rotation=25, ha='right', fontsize=7.5)
        ax.set_ylabel("Avg Rating", color=C_SUB, fontsize=8)
        ax.set_title("Ratings — Current vs Previous", color=C_TEXT, fontsize=10, fontweight='bold', pad=8)
        ax.set_ylim(3.0, 5.0)
        ax.legend(fontsize=7, facecolor=C_PANEL, edgecolor='#333', labelcolor=C_TEXT)
        fig.tight_layout(pad=1.2)
        return fig_to_rl(fig, 175, 60)

    # ── Build PDF ─────────────────────────────────────────────────────────────
    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                              topMargin=12*mm, bottomMargin=12*mm,
                              leftMargin=12*mm, rightMargin=12*mm)
    rl   = getSampleStyleSheet()
    dark = colors.HexColor("#0a0a0a")
    gold = colors.HexColor("#e8a020")
    story = []

    def sh(text):
        return Paragraph(text, ParagraphStyle('sh', parent=rl['Normal'],
            fontSize=11, textColor=gold, fontName='Helvetica-Bold', spaceBefore=6, spaceAfter=3))
    def gr():
        return HRFlowable(width="100%", thickness=1, color=gold, spaceAfter=4)
    def cap(text):
        return Paragraph(text, ParagraphStyle('cap', parent=rl['Normal'],
            fontSize=8, textColor=colors.HexColor("#555555"), spaceAfter=5))

    # Page 1: Header + summary table
    story.append(Paragraph(f"RollsKing — {report_type} Operations Report",
        ParagraphStyle('t', parent=rl['Title'], fontSize=18,
                       textColor=gold, backColor=dark, alignment=TA_CENTER, spaceAfter=2)))
    story.append(Paragraph(
        f"{label_curr} vs {label_prev}  |  Generated {datetime.now().strftime('%d %b %Y, %H:%M')}  |  "
        f"{len(summaries_curr)} Cities",
        ParagraphStyle('sub', parent=rl['Normal'], fontSize=9,
                       textColor=colors.HexColor("#666"), alignment=TA_CENTER, spaceAfter=8)))
    story.append(gr())

    # Summary table
    tbl_data = [["City / Manager", "Outlets",
                 f"Orders\n{label_curr}", f"Orders\n{label_prev}",
                 f"Cmp%\n{label_curr}", f"Cmp%\n{label_prev}",
                 f"KPT\n{label_curr}", f"KPT\n{label_prev}", "Breaches",
                 f"Rating\n{label_curr}", f"Rating\n{label_prev}"]]

    tier_rows = {}
    for idx, city in enumerate(summaries_curr, 1):
        curr = summaries_curr[city]
        prev = summaries_prev.get(city, {})
        row = [
            city.split('(')[0].strip(),
            curr.get('outlets','—'),
            curr.get('total_orders','—'), prev.get('total_orders','—'),
            f"{curr['cmp_pct']:.2f}%" if curr.get('cmp_pct') else "—",
            f"{prev['cmp_pct']:.2f}%" if prev.get('cmp_pct') else "—",
            f"{curr['avg_kpt']:.1f}" if curr.get('avg_kpt') else "—",
            f"{prev['avg_kpt']:.1f}" if prev.get('avg_kpt') else "—",
            curr.get('kpt_breaches','—'),
            f"{curr['avg_rating']:.2f}" if curr.get('avg_rating') else "—",
            f"{prev['avg_rating']:.2f}" if prev.get('avg_rating') else "—",
        ]
        tbl_data.append(row)
        # Flag high complaint rows
        if curr.get('cmp_pct') and curr['cmp_pct'] >= 3:
            tier_rows[idx] = 'red'
        elif curr.get('avg_kpt') and curr['avg_kpt'] > 12:
            tier_rows[idx] = 'amber'

    col_w = [38*mm,13*mm,16*mm,16*mm,14*mm,14*mm,13*mm,13*mm,13*mm,14*mm,14*mm]
    tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    ts  = TableStyle([
        ('BACKGROUND', (0,0),(-1,0), dark),
        ('TEXTCOLOR',  (0,0),(-1,0), gold),
        ('FONTNAME',   (0,0),(-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0),(-1,-1), 7.5),
        ('ALIGN',      (0,0),(-1,-1), 'CENTER'),
        ('ALIGN',      (0,1),(0,-1),  'LEFT'),
        ('ROWBACKGROUNDS', (0,1),(-1,-1), [colors.HexColor("#0f0f0f"), colors.HexColor("#141414")]),
        ('TEXTCOLOR',  (0,1),(-1,-1), colors.HexColor("#f0f0f0")),
        ('GRID',       (0,0),(-1,-1), 0.4, colors.HexColor("#2a2a2a")),
        ('VALIGN',     (0,0),(-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0),(-1,-1), 3),
        ('BOTTOMPADDING',(0,0),(-1,-1), 3),
    ])
    for row_idx, flag in tier_rows.items():
        clr = colors.HexColor("#2a0a0a") if flag == 'red' else colors.HexColor("#2a1a00")
        ts.add('BACKGROUND', (0,row_idx), (-1,row_idx), clr)
    tbl.setStyle(ts)
    story.append(tbl)

    # Page 2: Charts
    story.append(PageBreak())
    story.append(sh("Complaints & KPT — City Comparison")); story.append(gr())
    story.append(cap(f"Left: Complaint % per city. Red dashed = 3% threshold. Right: Avg KPT. Red dashed = 12 min threshold."))
    story.append(chart_complaints_kpt())

    rat_chart = chart_ratings()
    if rat_chart:
        story.append(Spacer(1, 6*mm))
        story.append(sh("Ratings — City Comparison")); story.append(gr())
        story.append(cap("Red dashed line = 4.0 minimum threshold. Red bars = below threshold."))
        story.append(rat_chart)

    doc.build(story)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for k, v in [('logged_in', False), ('dl_excel', None), ('dl_pdf', None),
              ('dl_excel_name', None), ('dl_pdf_name', None)]:
    if k not in st.session_state: st.session_state[k] = v

# ══════════════════════════════════════════════════════════════════════════════
# LOGIN
# ══════════════════════════════════════════════════════════════════════════════
APP_PASSWORD = "rollsking2025"
if not st.session_state.logged_in:
    col1, col2, col3 = st.columns([1,1,1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown('<div class="main-title">RK Ops</div>', unsafe_allow_html=True)
        st.markdown('<div class="main-sub">Daily & Weekly Operations Monitor</div>', unsafe_allow_html=True)
        pw = st.text_input("Password", type="password", placeholder="Enter password")
        if st.button("Sign In", use_container_width=True):
            if pw == APP_PASSWORD:
                st.session_state.logged_in = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="main-title">⚡ RK Ops Monitor</div>', unsafe_allow_html=True)
st.markdown('<div class="main-sub">Daily & Weekly Operations — Complaints · KPT · Ratings · Orders</div>',
            unsafe_allow_html=True)

tab_daily, tab_weekly = st.tabs(["📅  Daily Report", "📆  Weekly Report"])

def render_report_tab(report_type, label_curr_default, label_prev_default,
                       file_label_curr, file_label_prev, tab_key):

    col_upload, col_config = st.columns([1.2, 0.8])

    with col_upload:
        st.markdown(f"""<div style="margin-bottom:0.4rem;">
            <span class="step-badge">1</span>
            <span style="font-family:'Syne',sans-serif;font-size:0.9rem;font-weight:700;color:#fff;">Upload Data Files</span>
        </div>""", unsafe_allow_html=True)

        st.markdown(f'<div class="section-label">{file_label_curr}</div>', unsafe_allow_html=True)
        f_curr = st.file_uploader(f"Current {report_type} file",
                                   type=["xlsx"], key=f"curr_{tab_key}",
                                   label_visibility="collapsed")

        st.markdown(f'<div class="section-label" style="margin-top:0.6rem;">{file_label_prev}</div>',
                    unsafe_allow_html=True)
        f_prev = st.file_uploader(f"Previous {report_type} file",
                                   type=["xlsx"], key=f"prev_{tab_key}",
                                   label_visibility="collapsed")

        if f_curr:
            ok, _ = detect_file(f_curr.read()); f_curr.seek(0)
            st.markdown(f'<span class="{"chip-green" if ok else "chip-red"}">{"✓" if ok else "⚠"} {f_curr.name} — {"Recognised" if ok else "No Zomato/Swiggy sheet found"}</span>',
                        unsafe_allow_html=True)
        if f_prev:
            ok, _ = detect_file(f_prev.read()); f_prev.seek(0)
            st.markdown(f'<span class="{"chip-green" if ok else "chip-red"}">{"✓" if ok else "⚠"} {f_prev.name} — {"Recognised" if ok else "No Zomato/Swiggy sheet found"}</span>',
                        unsafe_allow_html=True)

    with col_config:
        st.markdown(f"""<div style="margin-bottom:0.4rem;">
            <span class="step-badge">2</span>
            <span style="font-family:'Syne',sans-serif;font-size:0.9rem;font-weight:700;color:#fff;">Configure</span>
        </div>""", unsafe_allow_html=True)

        label_curr = st.text_input("Label for Current Period", value=label_curr_default,
                                    key=f"lc_{tab_key}")
        label_prev = st.text_input("Label for Previous Period", value=label_prev_default,
                                    key=f"lp_{tab_key}")

        st.markdown("<div style='margin:0.5rem 0;'></div>", unsafe_allow_html=True)
        cities_all   = list(CITY_MAPPING.keys())
        cities_sel   = st.multiselect("Filter cities (leave blank = all)",
                                       options=cities_all, default=[],
                                       key=f"cities_{tab_key}")
        active_cities = cities_sel if cities_sel else cities_all

    st.markdown("<hr class='div'>", unsafe_allow_html=True)

    st.markdown(f"""<div style="margin-bottom:0.6rem;">
        <span class="step-badge">3</span>
        <span style="font-family:'Syne',sans-serif;font-size:0.9rem;font-weight:700;color:#fff;">Generate Report</span>
    </div>""", unsafe_allow_html=True)

    if not f_curr:
        st.markdown("""<div style="background:#111;border:1px dashed #2a2a2a;border-radius:8px;
        padding:1rem;color:#444;font-size:0.85rem;text-align:center;">
            Upload the current period file above to continue
        </div>""", unsafe_allow_html=True)
        return

    if st.button(f"⚡ Generate {report_type} Report", key=f"gen_{tab_key}"):
        with st.spinner("Processing..."):
            try:
                # Load current
                wb_curr = openpyxl.load_workbook(io.BytesIO(f_curr.read()), data_only=True)
                zmt_curr = load_zomato(wb_curr)
                swg_curr = load_swiggy(wb_curr)

                # Load previous
                zmt_prev = {}; swg_prev = {}
                if f_prev:
                    wb_prev = openpyxl.load_workbook(io.BytesIO(f_prev.read()), data_only=True)
                    zmt_prev = load_zomato(wb_prev)
                    swg_prev = load_swiggy(wb_prev)

                # Calculate per city
                summaries_curr = {}; summaries_prev = {}
                outlet_detail  = {}

                for city in active_cities:
                    outlets = CITY_MAPPING[city]
                    summ, rows = calc_city_summary(city, outlets, zmt_curr, swg_curr)
                    if summ:
                        summaries_curr[city] = summ
                        outlet_detail[city]  = rows

                    if f_prev:
                        summ_p, rows_p = calc_city_summary(city, outlets, zmt_prev, swg_prev)
                        if summ_p:
                            summaries_prev[city] = summ_p
                            outlet_detail[f"_prev_{city}"] = rows_p

                if not summaries_curr:
                    st.error("No matching outlet data found in the file. Check that IDs are correct.")
                    return

                # Live preview on screen
                st.markdown(f'<div class="section-label" style="margin-top:0.8rem;">Results — {label_curr} vs {label_prev}</div>',
                            unsafe_allow_html=True)

                # Metrics cards
                cols = st.columns(4)
                total_orders = sum(s['total_orders'] for s in summaries_curr.values())
                total_comps  = sum(s['total_comps']  for s in summaries_curr.values())
                all_cmp_pct  = round(total_comps/total_orders*100, 2) if total_orders else 0
                kpt_vals     = [s['avg_kpt'] for s in summaries_curr.values() if s.get('avg_kpt')]
                avg_kpt      = round(sum(kpt_vals)/len(kpt_vals), 2) if kpt_vals else 0
                rat_vals     = [s['avg_rating'] for s in summaries_curr.values() if s.get('avg_rating')]
                avg_rating   = round(sum(rat_vals)/len(rat_vals), 2) if rat_vals else 0
                kpt_breach   = sum(s.get('kpt_breaches',0) for s in summaries_curr.values())

                with cols[0]: st.metric("Total Orders",   f"{total_orders:,}")
                with cols[1]:
                    cmp_delta = None
                    if summaries_prev:
                        po = sum(s.get('total_orders',0) for s in summaries_prev.values())
                        pc = sum(s.get('total_comps',0)  for s in summaries_prev.values())
                        prev_pct = round(pc/po*100, 2) if po else 0
                        cmp_delta = round(all_cmp_pct - prev_pct, 2)
                    st.metric("Complaint %", f"{all_cmp_pct:.2f}%",
                              delta=f"{cmp_delta:+.2f}pp" if cmp_delta is not None else None,
                              delta_color="inverse")
                with cols[2]:
                    st.metric("Avg KPT", f"{avg_kpt:.1f} min",
                              delta=f"{kpt_breach} breaches", delta_color="inverse")
                with cols[3]: st.metric("Avg Rating", f"{avg_rating:.2f}" if avg_rating else "—")

                st.markdown("<div style='margin:0.5rem 0;'></div>", unsafe_allow_html=True)

                # City table
                table_rows = []
                for city, curr in summaries_curr.items():
                    prev = summaries_prev.get(city, {})
                    d_cmp = round(curr['cmp_pct'] - prev['cmp_pct'], 2) if (curr.get('cmp_pct') and prev.get('cmp_pct')) else None
                    cmp_flag = "🔴" if (curr.get('cmp_pct') and curr['cmp_pct'] >= 3) else ("🟡" if curr.get('cmp_pct') and curr['cmp_pct'] >= 2 else "🟢")
                    kpt_flag = "🔴" if curr.get('kpt_breaches') and curr['kpt_breaches'] > 0 else "🟢"
                    rat_flag = "🔴" if (curr.get('avg_rating') and curr['avg_rating'] < 4.0) else "🟢"
                    table_rows.append({
                        "City": city.split('(')[0].strip(),
                        "Orders": curr.get('total_orders','—'),
                        f"Cmp% {label_curr}": f"{cmp_flag} {curr['cmp_pct']:.2f}%" if curr.get('cmp_pct') else "—",
                        "Cmp Δ": f"{d_cmp:+.2f}pp" if d_cmp is not None else "—",
                        f"KPT {label_curr}": f"{kpt_flag} {curr['avg_kpt']:.1f}m" if curr.get('avg_kpt') else "—",
                        "KPT Breaches": curr.get('kpt_breaches','—'),
                        f"Rating {label_curr}": f"{rat_flag} {curr['avg_rating']:.2f}" if curr.get('avg_rating') else "—",
                    })

                import pandas as pd
                st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

                # Flagged outlets
                flagged = []
                for city, rows in outlet_detail.items():
                    if city.startswith('_prev_'): continue
                    for r in rows:
                        issues = []
                        if r.get('cmp_pct') and r['cmp_pct'] >= 3:
                            issues.append(f"Cmp {r['cmp_pct']:.1f}%")
                        if r.get('kpt') and r['kpt'] > 12:
                            issues.append(f"KPT {r['kpt']:.1f}m")
                        if r.get('rating') and r['rating'] < 4.0:
                            issues.append(f"Rating {r['rating']:.2f}")
                        if issues:
                            flagged.append({"City": city.split('(')[0].strip(),
                                            "Outlet": r['outlet'],
                                            "Issues": " · ".join(issues)})

                if flagged:
                    st.markdown('<div class="section-label" style="margin-top:0.8rem;color:#f87171;">🔴 Flagged Outlets</div>',
                                unsafe_allow_html=True)
                    st.dataframe(pd.DataFrame(flagged), use_container_width=True, hide_index=True)

                # Build downloads
                excel_bytes = build_excel_report(
                    summaries_curr, summaries_prev, outlet_detail,
                    report_type, label_curr, label_prev)
                pdf_bytes = build_pdf_report(
                    summaries_curr, summaries_prev, outlet_detail,
                    report_type, label_curr, label_prev)

                slug = f"RK_{report_type.replace(' ','_')}_{label_curr.replace(' ','_').replace('/','_')}"
                st.session_state.dl_excel      = excel_bytes
                st.session_state.dl_pdf        = pdf_bytes
                st.session_state.dl_excel_name = f"{slug}.xlsx"
                st.session_state.dl_pdf_name   = f"{slug}.pdf"

                st.success(f"✓ Report ready — {len(summaries_curr)} cities · {len(flagged)} flagged outlets")

            except Exception as e:
                import traceback
                st.error(f"Error: {e}")
                st.code(traceback.format_exc())

    # Downloads (persist across reruns)
    if st.session_state.dl_excel:
        st.markdown("<hr class='div'>", unsafe_allow_html=True)
        st.markdown('<div class="section-label">Download</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            st.download_button("📥 Download Excel",
                data=st.session_state.dl_excel,
                file_name=st.session_state.dl_excel_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_xlsx_{tab_key}")
        with c2:
            st.download_button("📄 Download PDF",
                data=st.session_state.dl_pdf,
                file_name=st.session_state.dl_pdf_name,
                mime="application/pdf",
                key=f"dl_pdf_{tab_key}")

# ── TABS ─────────────────────────────────────────────────────────────────────
with tab_daily:
    today     = date.today()
    yesterday = today - timedelta(days=1)
    render_report_tab(
        report_type       = "Daily",
        label_curr_default= today.strftime("%d %b %Y"),
        label_prev_default= yesterday.strftime("%d %b %Y"),
        file_label_curr   = "TODAY'S FILE — Zomato + Swiggy daily export",
        file_label_prev   = "YESTERDAY'S FILE — for comparison",
        tab_key           = "daily"
    )

with tab_weekly:
    week_start = today - timedelta(days=today.weekday())
    prev_week  = week_start - timedelta(days=7)
    render_report_tab(
        report_type       = "Weekly",
        label_curr_default= f"W/E {(week_start + timedelta(days=6)).strftime('%d %b')}",
        label_prev_default= f"W/E {(prev_week + timedelta(days=6)).strftime('%d %b')}",
        file_label_curr   = "THIS WEEK'S FILE — Zomato + Swiggy weekly export",
        file_label_prev   = "LAST WEEK'S FILE — for comparison",
        tab_key           = "weekly"
    )

st.markdown("""
<div style='text-align:center;color:#1e1e1e;font-size:0.75rem;padding:2rem 0 0.5rem;'>
    RollsKing Operations — City Leads & Zonal Head Monitor
</div>
""", unsafe_allow_html=True)
